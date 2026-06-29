import json, os, sys, tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="4F7FFF")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=15, color="1A2340")
BODY = Font(name=FONT, size=10, color="1A2340")
LOWF = Font(name=FONT, size=10, color="C62828")
thin = Side(style="thin", color="D9DEEC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")

res_path = os.environ.get("BULK_JSON") or os.path.join(tempfile.gettempdir(), "saemwork_bulk_results.json")
if not os.path.exists(res_path):
    alt = os.path.expandvars(r"%LOCALAPPDATA%\Temp\saemwork_bulk_results.json")
    if os.path.exists(alt):
        res_path = alt
with open(res_path, encoding="utf-8") as f:
    data = json.load(f)
summary, rows = data["summary"], data["rows"]

src = "쌤워크_피드백_서식_v2.xlsx"
out = sys.argv[1] if len(sys.argv) > 1 else "쌤워크_피드백_서식_v3.xlsx"
try:
    wb = load_workbook(src)
except Exception:
    wb = Workbook()  # 원본이 없으면 새로 만든다

name = f"자동검수결과({summary['n']}건)"
for s in [s for s in wb.sheetnames if s.startswith("자동검수결과")]:
    del wb[s]
ws = wb.create_sheet(name)
ws.sheet_view.showGridLines = False

ws["A1"] = "자동 검수 결과 — 비식별 샘플 500건 (규칙 엔진 자연스러움/품질)"
ws["A1"].font = TITLE_FONT
cat = " · ".join(f"{k} {v}" for k, v in summary["catAvg"].items())
lines = [
    f"검사 일시: {summary['generatedAt']}",
    f"표본 수: {summary['n']}건  ·  점검 방식: processRecord 생성 → qualityScorer 채점(0~100)",
    f"관찰일지 평균 {summary['obsAvg']} (최저 {summary['obsMin']}, 최고 {summary['obsMax']})",
    f"알림장 평균 {summary['noticeAvg']} (최저 {summary['noticeMin']}, 최고 {summary['noticeMax']})",
    f"아이 발화 보존: {summary['speechKept']}/{summary['quoted']} (따옴표 포함 입력)  ·  70점 미만 관찰일지: {summary['lowCount']}건",
    f"영역별 관찰 평균: {cat}",
]
for i, t in enumerate(lines, start=2):
    ws.cell(row=i, column=1, value=t).font = Font(name=FONT, size=11, bold=(i in (4, 5)))

hr = 9
headers = ["번호", "영역", "이름", "입력 메모(비식별)", "생성 — 관찰일지", "생성 — 알림장", "관찰점수", "알림장점수", "발화보존"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CTR; cell.border = BORDER
ws.freeze_panes = ws.cell(row=hr + 1, column=1)

for j, r in enumerate(rows):
    rr = hr + 1 + j
    speech = "—" if not r["hasSpeech"] else ("보존" if r["speechKept"] else "손실")
    vals = [r["idx"], r["category"], r["childName"], r["memo"], r["observation"], r["parent"],
            r["obsScore"], r["noticeScore"], speech]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=rr, column=c, value=v)
        cell.border = BORDER
        cell.alignment = WRAP if c in (4, 5, 6) else CTR
        cell.font = LOWF if (c == 7 and r["obsScore"] < 70) else BODY

widths = [6, 10, 8, 34, 40, 40, 9, 10, 9]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w

# 새 시트를 맨 앞쪽(안내 다음)으로 이동
try:
    wb.move_sheet(name, -(len(wb.sheetnames) - 2))
except Exception:
    pass

wb.save(out)
print("saved", out, "| sheet:", name, "| rows:", len(rows))
