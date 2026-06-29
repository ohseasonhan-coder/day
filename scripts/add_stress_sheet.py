import json, os, sys, tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="4F7FFF")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE = Font(name=FONT, bold=True, size=15, color="1A2340")
BODY = Font(name=FONT, size=11, color="1A2340")
thin = Side(style="thin", color="D9DEEC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")

res = os.environ.get("STRESS_JSON")
if not res or not os.path.exists(res):
    cand = [os.path.join(tempfile.gettempdir(), "saemwork_stress_results.json"),
            os.path.expandvars(r"%LOCALAPPDATA%\Temp\saemwork_stress_results.json")]
    res = next((c for c in cand if os.path.exists(c)), cand[0])
with open(res, encoding="utf-8") as f:
    data = json.load(f)
s, lowest = data["summary"], data.get("lowest", [])

src = "쌤워크_피드백_서식_v3.xlsx"
out = sys.argv[1] if len(sys.argv) > 1 else src
try:
    wb = load_workbook(src)
except Exception:
    wb = Workbook()

name = f"스트레스({s['n']//1000}천건)요약" if s['n'] < 100000 else f"스트레스({s['n']//10000}만건)요약"
for x in [x for x in wb.sheetnames if x.startswith("스트레스")]:
    del wb[x]
ws = wb.create_sheet(name)
ws.sheet_view.showGridLines = False

ws["A1"] = f"대량 스트레스/안정성 점검 — {s['n']:,}건 (규칙 엔진 관찰일지)"
ws["A1"].font = TITLE
rows = [
    ("검사 일시", s["generatedAt"]),
    ("처리 건수", f"{s['n']:,}건"),
    ("오류 / 빈결과", f"{s['errors']} / {s['empty']}"),
    ("소요 시간", f"{s['elapsedMs']/1000:.1f}초 (건당 {s['perRecordMs']}ms)"),
    ("관찰일지 평균", f"{s['obsAvg']} (최저 {s['obsMin']}, 최고 {s['obsMax']})"),
    ("아이 발화 보존", f"{s['speechKept']:,} / {s['quoted']:,}"),
]
r = 3
for k, v in rows:
    ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=11)
    ws.cell(row=r, column=2, value=v).font = BODY
    r += 1

r += 1
ws.cell(row=r, column=1, value="점수 분포").font = Font(name=FONT, bold=True, size=12); r += 1
for c, h in enumerate(["구간", "건수", "비율"], start=1):
    cell = ws.cell(row=r, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CTR; cell.border = BORDER
r += 1
for k, v in s["buckets"].items():
    ws.cell(row=r, column=1, value=k).border = BORDER
    ws.cell(row=r, column=2, value=v).border = BORDER
    ws.cell(row=r, column=3, value=f"{v/s['n']*100:.1f}%").border = BORDER
    r += 1

r += 1
ws.cell(row=r, column=1, value="영역별 관찰 평균").font = Font(name=FONT, bold=True, size=12); r += 1
for c, h in enumerate(["영역", "평균"], start=1):
    cell = ws.cell(row=r, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CTR; cell.border = BORDER
r += 1
for k, v in sorted(s["catAvg"].items(), key=lambda x: x[1]):
    ws.cell(row=r, column=1, value=k).border = BORDER
    ws.cell(row=r, column=2, value=v).border = BORDER
    r += 1

r += 1
ws.cell(row=r, column=1, value="최저 점수 예시(상위 15)").font = Font(name=FONT, bold=True, size=12); r += 1
for c, h in enumerate(["점수", "생성된 관찰일지"], start=1):
    cell = ws.cell(row=r, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CTR; cell.border = BORDER
r += 1
for item in lowest[:15]:
    ws.cell(row=r, column=1, value=item["score"]).border = BORDER
    cc = ws.cell(row=r, column=2, value=item["text"]); cc.border = BORDER; cc.alignment = WRAP
    r += 1

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 70
ws.column_dimensions["C"].width = 12
try:
    wb.move_sheet(name, -(len(wb.sheetnames) - 2))
except Exception:
    pass
wb.save(out)
print("saved", out, "| sheet:", name)
