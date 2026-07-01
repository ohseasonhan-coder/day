# v3 피드백 엑셀 분석·비식별 변환 도구 (읽기 전용).
#  - 원본 파일은 절대 수정하지 않는다(openpyxl read_only=True).
#  - 개인정보를 제외한 메타데이터 통계만 콘솔에 출력한다.
#  - 비식별 골든 데이터셋 초안은 data/golden_local/ (gitignore)로만 저장한다.
#  - 이름 컬럼 + 앱 합성 이름 풀(bulkSamples)까지 모두 비식별한다.
# 사용: python scripts/analyze_v3.py [원본경로] [출력경로]
import os, re, sys, json, statistics
from collections import Counter, defaultdict

# 앱 합성 이름 풀(src/utils/ai/datasets/bulkSamples.js NAMES). 목표 문장 본문에 남은 이름까지 비식별.
APP_NAMES = ['지우','서연','도윤','하준','수아','민준','예린','시우','하은','주아',
             '연우','지호','서윤','건우','아인','윤서','준서','다은','재윤','소율']

def load_app_names():
    try:
        p = 'src/utils/ai/datasets/bulkSamples.js'
        with open(p, encoding='utf-8') as f:
            m = re.search(r"const NAMES\s*=\s*\[([^\]]+)\]", f.read())
        if m:
            return re.findall(r"'([가-힣]{2,3})'", m.group(1))
    except Exception:
        pass
    return APP_NAMES

def parse_sections(text):
    """복사용(col7) → 3섹션. '관찰내용:' / '배움 읽기:' / '교사 지원 및 다음 계획:' 및 [대괄호] 모두 처리."""
    t = str(text or '').replace('\r', '')
    def grab(labels):
        for lb in labels:
            m = re.search(rf'(?:\[{lb}\]|{lb}\s*[:：])\s*(.*?)(?=(?:\n?\s*(?:\[(?:관찰내용|배움 읽기|교사 지원[^\]]*)\]|(?:관찰내용|배움 읽기|교사 지원[^:：]*)\s*[:：]))|$)', t, re.S)
            if m:
                return m.group(1).strip()
        return ''
    return {
        'observation': grab(['관찰내용']),
        'learning': grab(['배움 읽기', '배움읽기']),
        'support': grab(['교사 지원 및 다음 계획', '교사 지원', '지원 및 다음 계획']),
    }

def main():
    try: sys.stdout.reconfigure(encoding="utf-8")
    except Exception: pass
    src = sys.argv[1] if len(sys.argv) > 1 else "쌤워크_피드백_서식_v3.xlsx"
    out_dir = "data/golden_local"
    out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(out_dir, "observation_golden.local.json")
    if not os.path.exists(src):
        print(json.dumps({"error": f"원본 없음: {src}"}, ensure_ascii=False)); return

    from openpyxl import load_workbook
    wb = load_workbook(src, read_only=True, data_only=True)  # 읽기 전용
    target = next((s for s in wb.sheetnames if "자동검수" in s), wb.sheetnames[0])
    ws = wb[target]

    header = [ws.cell(9, c).value for c in range(1, ws.max_column + 1)]
    # 열(1-base): 3=이름, 4=입력, 5=관찰/배움, 6=교사지원, 7=복사용
    rows = []
    for r in ws.iter_rows(min_row=10):
        vals = [c.value for c in r]
        if not any(v not in (None, "") for v in vals):
            continue
        rows.append(vals)

    def cell(v): return "" if v is None else str(v).strip()
    name_col  = [cell(r[2]) if len(r) > 2 else "" for r in rows]
    input_col = [cell(r[3]) if len(r) > 3 else "" for r in rows]
    obs_col   = [cell(r[4]) if len(r) > 4 else "" for r in rows]
    copy_col  = [cell(r[6]) if len(r) > 6 else "" for r in rows]

    total = len(rows)
    blank_input = sum(1 for x in input_col if not x)
    dup_input = total - len(set(input_col))
    lens = [len(x) for x in input_col if x]
    length_dist = {
        "min": min(lens) if lens else 0, "max": max(lens) if lens else 0,
        "mean": round(statistics.mean(lens), 1) if lens else 0,
        "median": statistics.median(lens) if lens else 0,
    }
    linkable = sum(1 for i in range(total) if input_col[i] and obs_col[i])
    example_only = sum(1 for i in range(total) if (not input_col[i]) and copy_col[i])
    incomplete = sum(1 for i in range(total) if not input_col[i] and not copy_col[i])

    # ── 비식별 매핑 ──────────────────────────────────────────────────────
    all_names = list(dict.fromkeys([n for n in name_col if n] + load_app_names()))
    def anon_id(i):
        return chr(65 + i % 26) + ("" if i < 26 else str(i // 26 + 1)) + "원아"
    # 이름 컬럼 값 → 안정 라벨(A원아..), 앱 합성 이름 → 라벨 뒤에 이어서 매핑
    col_names = [n for n in dict.fromkeys(name_col) if n]
    name_map = {n: anon_id(i) for i, n in enumerate(col_names)}
    josa = r'(이|가|은|는|을|를|와|과|에게|의|도|만|께)?'
    # 앱 합성 이름은 조사까지 흡수해 '○○'로 redact(라벨 불일치·조사 오류 방지)
    pool_only = [n for n in load_app_names() if n not in name_map]
    def deidentify(text):
        t = text
        for real, anon in name_map.items():
            t = re.sub(re.escape(real) + josa, lambda m: anon + (m.group(1) or ''), t)
        for real in pool_only:
            t = re.sub(re.escape(real) + josa, '○○', t)
        return t

    # 잔여 실명 점검(비식별 검증)
    residual = Counter()
    for i in range(total):
        for nm in load_app_names():
            if re.search(re.escape(nm), deidentify(copy_col[i])):
                residual[nm] += 1

    # ── 중복/목표 통계(고유 입력 기준) ───────────────────────────────────
    def norm(t): return re.sub(r'[A-Z]원아|○○', '○', deidentify(t)).strip()
    groups = defaultdict(list)
    for i in range(total):
        groups[norm(input_col[i])].append(i)
    same_in_diff_out = sum(1 for g in groups.values() if len(g) > 1 and
                           len({norm(copy_col[j]) for j in g}) > 1)

    stats = {
        "sheet": target, "header": header, "rows": total,
        "blank_input_ratio": round(blank_input / total, 3) if total else 0,
        "duplicate_input_rows": dup_input,
        "unique_inputs": len(groups),
        "duplicate_groups": sum(1 for g in groups.values() if len(g) > 1),
        "same_input_different_target": same_in_diff_out,
        "input_length": length_dist,
        "classify": {"input_answer_linkable": linkable, "output_example_only": example_only, "incomplete": incomplete},
        "distinct_names_in_name_col": len(col_names),
        "residual_realname_after_deid": dict(residual),  # 비어 있어야 정상
    }
    print("=== v3 메타데이터 통계(개인정보 제외) ===")
    print(json.dumps(stats, ensure_ascii=False, indent=2))

    # ── 비식별 골든 초안 생성(목표 3섹션 포함) ──────────────────────────
    draft = {"generatedFrom": os.path.basename(src), "count": 0,
             "note": "비식별 로컬 전용 — Git 추적 제외. 목표(target)는 앱이 합성 이름으로 자동 생성한 참조 서식임.",
             "regressionCases": []}
    for i in range(total):
        if not input_col[i]:
            continue
        tgt = parse_sections(deidentify(copy_col[i]))
        draft["regressionCases"].append({
            "id": f"v3_local_{i+1:04d}",
            "documentType": "observation",
            "input": deidentify(input_col[i]),
            "factCard": {"name": name_map.get(name_col[i], "원아"),
                         "speeches": re.findall(r'"([^"]+)"', deidentify(input_col[i]))},
            "target": tgt,
            "reference": {"observationOrLearning": deidentify(obs_col[i]), "copyReady": deidentify(copy_col[i])},
        })
    draft["count"] = len(draft["regressionCases"])

    os.makedirs(out_dir, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(draft, f, ensure_ascii=False, indent=1)
    ok = "잔여 실명 0 ✅" if not residual else f"⚠ 잔여 실명 {sum(residual.values())}건"
    print(f"\n비식별 골든 초안 저장(로컬·gitignore): {out_path} · {draft['count']}건 · 비식별 {ok}")
    print("주의: 원본 엑셀은 수정하지 않았고, 로컬 초안은 Git 추적 제외 경로에만 저장됩니다.")

if __name__ == "__main__":
    main()
