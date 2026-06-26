from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="4F7FFF")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
EX_FILL = PatternFill("solid", fgColor="FFF7E6")
EX_FONT = Font(name=FONT, italic=True, color="8A5A00", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1A2340")
WARN_FONT = Font(name=FONT, bold=True, color="C62828", size=11)
BODY_FONT = Font(name=FONT, size=11, color="1A2340")
thin = Side(style="thin", color="D9DEEC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CTR
        cell.border = BORDER
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def add_dv(ws, options, col_letter, first=2, last=200):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(options), allow_blank=True)
    dv.error = "목록에서 선택해 주세요"
    dv.prompt = "선택: " + " / ".join(options)
    ws.add_data_validation(dv)
    dv.add("%s%d:%s%d" % (col_letter, first, col_letter, last))


wb = Workbook()

# ── 안내 ───────────────────────────────────────────────
g = wb.active
g.title = "안내"
g.sheet_view.showGridLines = False
g["A1"] = "쌤워크 피드백 / 검수 서식"
g["A1"].font = TITLE_FONT
g["A3"] = "⚠️ 실제 원아 이름·사진·주소·건강·가족 정보는 넣지 마세요. A원아 / ○○반 같은 비식별 예시로 적어주세요."
g["A3"].font = WARN_FONT
notes = [
    "",
    "사용법",
    "1) '문제보고' 시트: 앱이 이상하거나 불편할 때 한 줄씩 적습니다.",
    "2) '문장검수' 시트: 생성된 문장이 어색할 때 입력 메모와 결과를 적고 점수를 매깁니다.",
    "   - '원하는 결과' 칸에 직접 고쳐 쓴 문장을 적어주시면, 어디가 왜 어긋났는지 더 정확히 봐드립니다.",
    "3) 다 채우면 이 파일을 그대로 보내주세요. (해당되는 칸만 채워도 됩니다)",
    "",
    "점수 기준(문장검수): 0=미흡 · 1=보통 · 2=우수  /  민원위험은 '있음'이면 문제",
    "등급(문제보고): Blocker=즉시수정 · Major=베타중수정 · Minor=베타후 · Idea=아이디어",
]
r = 5
for line in notes:
    g.cell(row=r, column=1, value=line).font = (Font(name=FONT, bold=True, size=12) if line in ("사용법",) else BODY_FONT)
    r += 1
g.column_dimensions["A"].width = 95

# ── 문제보고 ────────────────────────────────────────────
s1 = wb.create_sheet("문제보고")
h1 = ["번호", "날짜", "화면", "무엇을 했나(버튼/입력)", "발생한 일", "기대한 결과",
      "기기", "브라우저", "재현빈도", "등급", "캡처", "상태"]
s1.append(h1)
ex1 = ["예시", "2026-06-23", "오늘기록", "알림장 카드 '복사' 누름 → 메모장 붙여넣기",
       "맨 앞에 빈 줄이 생김", "빈 줄 없이 바로 시작", "아이폰", "사파리", "항상", "Minor", "N", "접수"]
s1.append(ex1)
for c in range(1, len(h1) + 1):
    s1.cell(row=2, column=c).fill = EX_FILL
    s1.cell(row=2, column=c).font = EX_FONT
    s1.cell(row=2, column=c).alignment = WRAP
    s1.cell(row=2, column=c).border = BORDER
style_header(s1, len(h1))
widths1 = [6, 12, 12, 30, 26, 24, 12, 12, 10, 10, 7, 10]
for i, w in enumerate(widths1, start=1):
    s1.column_dimensions[chr(64 + i)].width = w
add_dv(s1, ["오늘기록", "문서함", "원아기록", "설정", "동기화", "알림장", "통계", "기타"], "C")
add_dv(s1, ["항상", "가끔", "한 번"], "I")
add_dv(s1, ["Blocker", "Major", "Minor", "Idea"], "J")
add_dv(s1, ["Y", "N"], "K")
add_dv(s1, ["접수", "수정중", "완료", "보류"], "L")
# 빈 입력 행 테두리
for row in range(3, 60):
    for c in range(1, len(h1) + 1):
        cell = s1.cell(row=row, column=c)
        cell.border = BORDER
        cell.font = BODY_FONT
        cell.alignment = WRAP

# ── 문장검수 ────────────────────────────────────────────
s2 = wb.create_sheet("문장검수")
h2 = ["번호", "문서유형", "입력 메모(비식별)", "생성된 문장(비식별)", "원하는 결과(직접 고쳐쓴 문장)",
      "사실보존", "발화보존", "자연스러움", "민원위험", "목적적합", "반복없음", "바로복사가능",
      "총평", "메모(왜 그렇게 바라는지)"]
s2.append(h2)
ex2 = ["예시", "관찰일지",
       'A원아가 블록으로 탑을 쌓다가 무너지자 "다시!"라고 말하며 다시 쌓았다.',
       "(앱이 만들어준 문장을 붙여넣기)",
       'A원아는 블록 탑이 무너졌지만 "다시!"라고 말하며 끝까지 다시 쌓는 끈기를 보였다.',
       2, 2, 2, "없음", 2, 2, 2, "수정필요", "결과가 너무 건조해서 끈기·태도가 드러나면 좋겠어요"]
s2.append(ex2)
for c in range(1, len(h2) + 1):
    s2.cell(row=2, column=c).fill = EX_FILL
    s2.cell(row=2, column=c).font = EX_FONT
    s2.cell(row=2, column=c).alignment = WRAP
    s2.cell(row=2, column=c).border = BORDER
style_header(s2, len(h2))
widths2 = [6, 14, 32, 32, 34, 9, 9, 10, 9, 9, 9, 11, 10, 28]
for i, w in enumerate(widths2, start=1):
    s2.column_dimensions[chr(64 + i)].width = w
add_dv(s2, ["관찰일지", "알림장", "보육일지 평가", "상담자료", "발달평가"], "B")
for col in ["F", "G", "H", "J", "K", "L"]:
    add_dv(s2, ["0", "1", "2"], col)
add_dv(s2, ["있음", "없음"], "I")
add_dv(s2, ["합격", "수정필요"], "M")
for row in range(3, 60):
    for c in range(1, len(h2) + 1):
        cell = s2.cell(row=row, column=c)
        cell.border = BORDER
        cell.font = BODY_FONT
        cell.alignment = WRAP

import sys
out = sys.argv[1] if len(sys.argv) > 1 else "쌤워크_피드백_서식.xlsx"
wb.save(out)
print("saved", out)
